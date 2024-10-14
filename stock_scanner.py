#!/usr/bin/env python
# coding: utf-8

# In[9]:


import yfinance as yf
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# Function to calculate the return over a specified period
def calculate_return(stock_symbol, period):
    stock_data = yf.Ticker(stock_symbol)
    hist = stock_data.history(period=period)

    if len(hist) > 0:
        start_price = hist['Close'].iloc[0]
        end_price = hist['Close'].iloc[-1]
        return_percentage = (end_price - start_price) / start_price * 100
        return start_price, end_price, return_percentage
    else:
        return None, None, None

# Function to scan stocks and save results
def daily_scan(stock_symbols, period='7y', target_return=0):
    # Get current date and time
    scan_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    file_name = "stock_scan_results.xlsx"

    # Store results
    detailed_results = []
    zero_return_stocks = []

    # Scan each stock
    for stock in stock_symbols:
        print(f"Scanning {stock}...")
        start_price, end_price, stock_return = calculate_return(stock, period)

        if start_price is not None:
            # Append detailed stock data to the list
            detailed_results.append([stock, start_price, end_price, stock_return])

            # Check if stock return meets the target return
            if abs(stock_return) <= target_return:
                zero_return_stocks.append([stock])

            # Print results in the console
            print(f"{stock}: Start Price = {start_price}, End Price = {end_price}, Return = {stock_return:.2f}%")
        else:
            print(f"Data for {stock} not available.")

    # Load existing Excel file or create a new one
    try:
        with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # Save detailed results to "Detailed Results" sheet
            if len(detailed_results) > 0:
                df_detailed = pd.DataFrame(detailed_results, columns=["Ticker", "Start Price", "End Price", "Return"])
                df_detailed['Scan Date'] = scan_time
                
                # Load existing workbook to find the last row
                book = load_workbook(file_name)
                if "Detailed Results" in book.sheetnames:
                    startrow = book["Detailed Results"].max_row  # Get max row for appending
                else:
                    startrow = 0
                
                df_detailed.to_excel(writer, sheet_name="Detailed Results", index=False, startrow=startrow)

            # Save 0% return stocks to "0% Return Stocks" sheet
            if len(zero_return_stocks) > 0:
                df_zero_return = pd.DataFrame(zero_return_stocks, columns=["Stocks with ~0% Return"])
            else:
                df_zero_return = pd.DataFrame([["None"]], columns=["Stocks with ~0% Return"])

            df_zero_return.insert(0, "Scan Date", scan_time)
            
            # Load existing workbook to find the last row for zero return stocks
            book = load_workbook(file_name)
            if "0% Return Stocks" in book.sheetnames:
                startrow = book["0% Return Stocks"].max_row  # Get max row for appending
            else:
                startrow = 0
                
            df_zero_return.to_excel(writer, sheet_name="0% Return Stocks", index=False, startrow=startrow)

        print(f"Results saved to {file_name}")

    except FileNotFoundError:
        # If file does not exist, create new file and save results
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            if len(detailed_results) > 0:
                df_detailed = pd.DataFrame(detailed_results, columns=["Ticker", "Start Price", "End Price", "Return"])
                df_detailed['Scan Date'] = scan_time
                df_detailed.to_excel(writer, sheet_name="Detailed Results", index=False)

            if len(zero_return_stocks) > 0:
                df_zero_return = pd.DataFrame(zero_return_stocks, columns=["Stocks with ~0% Return"])
            else:
                df_zero_return = pd.DataFrame([["None"]], columns=["Stocks with ~0% Return"])

            df_zero_return.insert(0, "Scan Date", scan_time)
            df_zero_return.to_excel(writer, sheet_name="0% Return Stocks", index=False)

        print(f"Created new Excel file and saved results to {file_name}")

# List of stock symbols to scan
stock_symbols = ['NVDA', 'AAPL', 'ADBE', 'META']  # Add your own list of stocks

# Run the scan
daily_scan(stock_symbols, period='10y', target_return=0)


# In[ ]:




